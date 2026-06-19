import openpyxl

wb = openpyxl.load_workbook('FINAL EAT_RIGHT_STATIONS DATA DRAFT - Copy.xlsx')
sheet = wb.active
print('Sheet title:', sheet.title)
print('Max row in openpyxl:', sheet.max_row)
print('Max col in openpyxl:', sheet.max_column)

# Count how many rows have any data
non_empty_rows = 0
for row_idx in range(1, sheet.max_row + 1):
    row_values = [sheet.cell(row_idx, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    is_empty = all(v is None for v in row_values)
    if not is_empty:
        non_empty_rows += 1
    else:
        print(f'Row {row_idx} is empty')

print('Total non-empty rows:', non_empty_rows)
